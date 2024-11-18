from tkinter import *
from tkinter import messagebox
import openpyxl
import pathlib
from matplotlib import pyplot as plt


root = Tk()
root.title("Data Entry")
root.geometry('700x400+300+200')
root.resizable(False, False)
root.configure(bg="#326273")



file=pathlib.Path("data.xlsx")
if file.exists():
    print("File exists")
    pass
else:
    file=openpyxl.Workbook()
    sheet=file.active
    sheet['A1']= "Date"
    sheet['B1']= "Calling Time(Mins.)"
    sheet['C1']= "Walking Steps"

    file.save("data.xlsx")
    
def submit_data():
    # Getting the values from entry fields
    date = dateValue.get()
    calling_time = callingValue.get()
    steps = stepsValue.get()

    file=openpyxl.load_workbook('data.xlsx')
    sheet = file.active
    sheet.cell(column=1,row=sheet.max_row,value=date)
    sheet.cell(column=2,row=sheet.max_row,value=calling_time)
    sheet.cell(column=3,row=sheet.max_row,value=steps)

    file.save(r'data.xlsx')

    # Check if any field is empty
    if not date or not steps or not calling_time:
        messagebox.showwarning("Input Error", "Please fill in all fields.")
    else:
        # Normally, you could save data to a file or database here
        messagebox.showinfo("Success", "Data Submitted Successfully")
        clear_fields()



# Icon
icon_image = PhotoImage(file="logo.png")  # Replace with the actual file path
root.iconphoto(False, icon_image)

# Title label
label_title = Label(root, text="Please fill the daily entry of the data:", font="Arial 13", bg="#326273", fg="#fff")
label_title.place(x=20, y=20)

# Labels for form fields
label_date = Label(root, text="Date", font="Arial 12", bg="#326273", fg="#fff")
label_date.place(x=50, y=60)


label_callingTime = Label(root, text="Calling Time", font="Arial 12", bg="#326273", fg="#fff")
label_callingTime.place(x=50, y=100)

label_steps = Label(root, text="Walking Steps", font="Arial 12", bg="#326273", fg="#fff")
label_steps.place(x=50, y=140)

# Entry fields
dateValue = StringVar()
callingValue = StringVar()
stepsValue = StringVar()

dateEntry = Entry(root, textvariable=dateValue, width=45, bd=2, font=20)
callingEntry = Entry(root, textvariable=callingValue, width=45, bd=2, font=20)
stepsEntry = Entry(root, textvariable=stepsValue, width=45, bd=2, font=20)

dateEntry.place(x=200, y=60)
callingEntry.place(x=200, y=100)
stepsEntry.place(x=200, y=140)

# Function to handle Submit action

# Function to clear all entry fields
def clear_fields():
    dateValue.set("")
    stepsValue.set("")
    callingValue.set("")

# Function to exit the application
def exit_application():
    root.quit()


def create_plot():
    workbook = openpyxl.load_workbook("data.xlsx")
    sheet = workbook.active  

    header = [cell.value for cell in sheet[1]]
    print("Header:", header)
    total_dates = []
    total_steps = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        date, calling_time, steps, _ = row
        if(date and steps):
            total_dates.append(date)
            total_steps.append(steps)
    plt.plot(total_dates, total_steps)
    plt.show()

        


submit_button = Button(root, text="Submit", command=submit_data, font="Arial 12", fg="#000", width=10)
submit_button.place(x=10, y=320)


clear_button = Button(root, text="Clear", command=clear_fields, font="Arial 12", fg="#000", width=10)
clear_button.place(x=160, y=320)

plot_button = Button(root, text="Plot", command=create_plot, font="Arial 12", fg="#000", width=10)
plot_button.place(x=310, y=320)


exit_button = Button(root, text="Exit", command=lambda:root.destroy(), font="Arial 12", fg="#000", width=10)
exit_button.place(x=460, y=320)

root.mainloop()
